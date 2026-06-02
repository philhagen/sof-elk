#!/usr/bin/env python3
# SOF-ELK® Supporting script
# (C)2026 Lewes Technology Consulting, LLC
#
# Converts a Hindsight (obsidianforensics) Internet History xlsx workbook
# into one-JSON-per-line, ready to drop into /logstash/kape/ for ingestion
# by SOF-ELK's 6508-browser_hindsight.conf parser.
#
# Hindsight is a Chromium-browser (Chrome/Edge/Brave/Opera) history parser.
# Its xlsx output has a "Timeline" sheet with TWO header rows:
#   row 1: section grouping (URL Specific / Download Specific / Cache Specific)
#   row 2: the actual field names
# This script skips row 1, uses row 2 as the header, and emits each data row
# as one JSON object with lowercased + sanitized keys.
#
# Field types are preserved (datetimes → ISO 8601 strings, ints → ints,
# floats → floats, etc.); empty cells are dropped.
#
# Usage:
#   hindsight2json.py -r INPUT.xlsx -w OUTPUT.json [-s SHEET] [-t TAG ...]
#
# Tip: Hindsight also has a native --format=jsonl output mode. If you have
# that, you can skip this converter entirely; the 6508 parser handles both.

import argparse
import json
import re
import sys
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    sys.stderr.write("ERROR: openpyxl is required (pip install openpyxl).\n")
    sys.exit(2)


def normalize_field_name(name):
    """Lowercase, spaces -> underscores, drop non-[a-z0-9_-]. Matches csv2json.py."""
    if not name:
        return ""
    fieldname = str(name).replace(" ", "_")
    fieldname = re.sub(r"[^a-zA-Z0-9\-_]", "", fieldname)
    return fieldname.lower()


def convert_value(value):
    """Normalize cell values: datetimes → ISO string, bool/int/float kept, strings stripped."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        # Hindsight timestamps in xlsx are already datetime objects; emit ISO 8601
        return value.isoformat(sep=" ")
    if isinstance(value, str):
        v = value.strip()
        if v == "":
            return None
        # Preserve numeric-like and boolean-like strings as proper types
        low = v.lower()
        if low == "true":
            return True
        if low == "false":
            return False
        try:
            return int(v)
        except ValueError:
            pass
        try:
            f = float(v)
            # avoid converting e.g. "1.5.3" to NaN — float() raises on those
            return f
        except ValueError:
            pass
        return v
    return value


def remove_empty(d):
    return {k: v for k, v in d.items() if v not in (None, "", [], {})}


def process_xlsx_to_jsonl(xlsx_filename, json_filename, sheet, tags):
    try:
        wb = openpyxl.load_workbook(xlsx_filename, read_only=True, data_only=True)
    except FileNotFoundError:
        sys.stderr.write(f"ERROR: File '{xlsx_filename}' not found.\n")
        sys.exit(1)

    if sheet not in wb.sheetnames:
        sys.stderr.write(
            f"ERROR: Sheet '{sheet}' not in workbook. Available: {wb.sheetnames}\n"
        )
        sys.exit(1)
    ws = wb[sheet]

    # Hindsight Timeline header is on row 2 (row 1 is section grouping).
    # For Storage / Extension Data sheets the header is on row 1.
    if sheet == "Timeline":
        header_row, data_start = 2, 3
    else:
        header_row, data_start = 1, 2

    headers = [normalize_field_name(c.value) for c in ws[header_row]]

    count = 0
    with open(json_filename, "w") as out:
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            obj = {}
            for k, v in zip(headers, row):
                if not k:
                    continue
                cv = convert_value(v)
                if cv is not None:
                    obj[k] = cv

            # Skip totally-empty rows (Hindsight occasionally writes blanks at the end)
            if not obj:
                continue

            if tags is not None:
                if "tags" in obj:
                    obj["tags"] = obj["tags"].split(" ") if isinstance(obj["tags"], str) else list(obj["tags"])
                else:
                    obj["tags"] = []
                for tag in tags:
                    obj["tags"].append(tag)

            json.dump(obj, out, default=str)
            out.write("\n")
            count += 1

    sys.stderr.write(
        f"hindsight2json: {count} rows  {xlsx_filename}[{sheet}] -> {json_filename}\n"
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert a Hindsight Internet History xlsx Timeline to JSONL."
    )
    parser.add_argument("-r", "--read", dest="infile", required=True,
                        help="Hindsight xlsx input file")
    parser.add_argument("-w", "--write", dest="outfile", required=True,
                        help="JSONL output file to create")
    parser.add_argument("-s", "--sheet", dest="sheet", default="Timeline",
                        help="Sheet to convert (default: Timeline)")
    parser.add_argument("-t", "--tag", dest="tags", action="append",
                        help='Optional string to add to "tags" field — repeatable')
    args = parser.parse_args()

    process_xlsx_to_jsonl(args.infile, args.outfile, args.sheet, args.tags)
